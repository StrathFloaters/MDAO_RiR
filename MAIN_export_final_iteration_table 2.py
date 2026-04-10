#!/usr/bin/env python
"""
WEIS SQL Log → Final-Iteration Spreadsheet
===========================================
Reads multiple WEIS optimisation SQL recorder files and outputs an Excel
spreadsheet (.xlsx) where:

  - Rows   = user-selected variables (objectives, design variables, constraints)
  - Columns = one per optimisation run, showing the value at the FINAL iteration

For **vector** variables (e.g. tower.diameter with multiple stations) the
script writes one row per component, labelled  "Variable [0]", "Variable [1]", …

Usage
-----
1. Edit the ``runs`` dictionary to point to your SQL files.
2. Edit ``select_objectives``, ``select_design_vars``, ``select_constraints``
   to choose which variables appear in the table.
3. Optionally adjust ``vector_reduce`` if you prefer a single summary row
   per vector variable instead of one row per component.
4. Run:  python MAIN_export_final_iteration.py

Requirements: openmdao, openpyxl, numpy
"""

import os
import numpy as np
import openmdao.api as om

# We use openpyxl via pandas, but keep a direct import for the
# optional formatting pass at the end.
try:
    import pandas as pd
except ImportError:
    raise SystemExit("pandas is required.  Install with:  pip install pandas")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl not found – the spreadsheet will be written as a plain "
          ".xlsx without formatting.  Install with:  pip install openpyxl")

# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  CONFIGURATION                                                          ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

# ---------------------------------------------------------------------------
# 1) Run dictionary  –  { 'Label': 'path/to/log_opt.sql', ... }
# ---------------------------------------------------------------------------
runs = {
    # 'A_twr': os.path.join(
        # "outputs",
        # "inp1_modA_anatowr",
        # "outputs",
        # "log_opt_inp1_modA_anatowr.sql",
    # ),
    # 'B_twr': os.path.join(
        # "outputs",
        # "inp1_modB_anatowr",
        # "outputs",
        # "log_opt_inp1_modB_anatowr.sql",
    # ),
    # 'C_twr': os.path.join(
        # "outputs",
        # "inp1_modC_anatowr",
        # "outputs",
        # "log_opt_inp1_modC_anatowr.sql",
    # ),
    'A_ptfm': os.path.join(
        "outputs",
        "inp1A_modA_anaptfm",
        "outputs",
        "log_opt_inp1A_modA_anaptfm.sql",
    ),
    'B_ptfm': os.path.join(
        "outputs",
        "inp1B_modB_anaptfm",
        "outputs",
        "log_opt_inp1B_modB_anaptfm.sql",
    ),
    'C_ptfm': os.path.join(
        "outputs",
        "inp1C_modC_anaptfm",
        "outputs",
        "log_opt_inp1C_modC_anaptfm.sql",
    ),
    'A_ptfm_twr': os.path.join(
        "outputs",
        "inp1_modA_anatower_ptfm",
        "outputs",
        "log_opt_inp1_modA_anatower_ptfm.sql",
    ),
    'B_ptfm_twr': os.path.join(
        "outputs",
        "inp1_modB_anatower_ptfm",
        "outputs",
        "log_opt_inp1_modB_anatower_ptfm.sql",
    ),
    'C_ptfm_twr': os.path.join(
        "outputs",
        "inp1_modC_anatower_ptfm",
        "outputs",
        "log_opt_inp1_modC_anatower_ptfm.sql",
    ),
}

# ---------------------------------------------------------------------------
# 2) Variables to include  –  substring matching (case-insensitive)
#    Set to None (or []) to include ALL variables of that category.
# ---------------------------------------------------------------------------
select_objectives  = [
    'floatingse.system_structural_mass',
    'towerse.tower_mass',
]
select_design_vars = [
    'floating.jointdv_0',
    'floating.jointdv_1',
    'floating.memgrp1.outer_diameter_in',
    'tower.diameter',
    'tower.layer_thickness',
]
select_constraints = [
    'raft.Max_PtfmPitch',
    'raft.max_nac_accel',
    'raft.Max_Offset',
    'floatingse.structural_frequencies',
]

# ---------------------------------------------------------------------------
# 3) Vector handling
#    'components' → one row per vector element  (e.g. tower.diameter[0], [1], …)
#    'mean'       → single row with np.mean of the vector
#    'max'        → single row with np.max
#    'min'        → single row with np.min
# ---------------------------------------------------------------------------
vector_mode = 'components'   # or 'mean', 'max', 'min'

# ---------------------------------------------------------------------------
# 4) Output filename
# ---------------------------------------------------------------------------
output_file = 'weis_final_iteration_comparison.xlsx'

# ---------------------------------------------------------------------------
# 5) Pretty-name / unit map  (same convention as the plotting script)
# ---------------------------------------------------------------------------
VAR_DISPLAY = {
    # Objectives
    'floatingse.system_structural_mass': ('Floating support structure mass', 'MTonne'),
    'financese.lcoe':                    ('LCOE',                           'USD/MWh'),
    'total_AEP':                         ('Annual Energy Production',       'GWh'),
    'towerse.tower_mass':                ('Tower structural mass',          'MTonne'),
    # Design variables
    'chord':                                   ('Blade Chord',                    'm'),
    'spar_cap_ss':                             ('Spar Cap SS Thickness',          'm'),
    'spar_cap_ps':                             ('Spar Cap PS Thickness',          'm'),
    'twist':                                   ('Blade Twist',                    'deg'),
    'member_diameter':                         ('Member Diameter',                'm'),
    'member_wall_thickness':                   ('Member Wall Thickness',          'm'),
    'joint_position':                          ('Joint Position',                 'm'),
    'floating.jointdv_0':                      ('Keel vert. pos. (draft)',        'm'),
    'floating.jointdv_1':                      ('Center to outer col. distance',  'm'),
    'floating.memgrp1.outer_diameter_in':      ('Outer column diameter',          'm'),
    'tower.diameter':                          ('Tower diameters',                'm'),
    'tower.thickness':                         ('Tower thicknesses',              'm'),
    # Constraints
    'tip_deflection':                       ('Tip Deflection Ratio',            ''),
    'freq_ratio':                           ('Frequency Ratio',                ''),
    'rotor_overspeed':                      ('Rotor Overspeed',                ''),
    'max_surge':                            ('Max Surge',                      'm'),
    'pitch_period':                         ('Pitch Period',                   's'),
    'heave_period':                         ('Heave Period',                   's'),
    'raft.Max_PtfmPitch':                   ('Max ptfm pitch (raft)',          'deg'),
    'raft.max_nac_accel':                   ('Max nacelle acceleration',       'm/s²'),
    'floatingse.structural_frequencies':    ('Eigenfrequencies, ptfm (raft)', 'Hz'),
    'floatingse.fore_aft_freqs':            ('1st F-A eigenfreq.', 'Hz'),
    'floatingse.side_side_freqs':           ('1st S-S eigenfreq.', 'Hz'),
    'towerse.post.constr_global_buckling':  ('Tower global buckling', ''),
    'towerse.post.constr_shell_buckling':   ('Tower (local) shell buckling', ''),
}


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  HELPER FUNCTIONS                                                       ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

def _get_display(var_name):
    """Return (pretty_name, unit_str) using substring matching in VAR_DISPLAY."""
    var_lower = var_name.lower()
    for key, (pretty, unit) in VAR_DISPLAY.items():
        if key.lower() in var_lower:
            return pretty, unit
    return var_name.split('.')[-1], ''


def read_weis_opt(sql_path):
    """Read optimisation history from a WEIS SQL recorder file."""
    cr = om.CaseReader(sql_path)
    cases = cr.list_cases('driver', recurse=False, out_stream=None)

    history = {'objectives': {}, 'design_vars': {}, 'constraints': {}}

    for case_name in cases:
        case = cr.get_case(case_name)
        for key, val in case.get_objectives().items():
            history['objectives'].setdefault(key, []).append(val.flatten())
        for key, val in case.get_design_vars().items():
            history['design_vars'].setdefault(key, []).append(val.flatten())
        for key, val in case.get_constraints().items():
            history['constraints'].setdefault(key, []).append(val.flatten())

    for category in history:
        for key in history[category]:
            history[category][key] = np.array(history[category][key])

    return history


def _match_vars(available, requested):
    """Substring-match requested names against available variable names."""
    if not requested:
        return available
    matched = []
    for var in available:
        var_lower = var.lower()
        for req in requested:
            if req.strip().lower() in var_lower:
                matched.append(var)
                break
    if not matched:
        print(f"  WARNING: none of {requested} matched available: {available}")
    return matched


def _reduce(arr, mode):
    """Reduce a 1-D array to a scalar using the chosen mode."""
    if mode == 'mean':
        return np.mean(arr)
    elif mode == 'max':
        return np.max(arr)
    elif mode == 'min':
        return np.min(arr)
    return np.mean(arr)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  MAIN                                                                    ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

def export_final_iteration(
    run_dict,
    select_objectives=None,
    select_design_vars=None,
    select_constraints=None,
    vector_mode='components',
    output_file='weis_final_iteration_comparison.xlsx',
):
    # ── 1. Read all SQL files ───────────────────────────────────────────
    all_histories = {}
    for label, path in run_dict.items():
        print(f"Reading {label} ← {path}")
        all_histories[label] = read_weis_opt(path)

    run_labels = list(run_dict.keys())

    # ── 2. Determine which variables to export ──────────────────────────
    ref = list(all_histories.values())[0]
    obj_vars = _match_vars(list(ref['objectives'].keys()),  select_objectives)
    dv_vars  = _match_vars(list(ref['design_vars'].keys()), select_design_vars)
    con_vars = _match_vars(list(ref['constraints'].keys()), select_constraints)

    # ── 3. Build rows ──────────────────────────────────────────────────
    # Each row: (category_label, pretty_name, unit, {run_label: value, ...})
    rows = []

    for cat_label, var_list, cat_key in [
        ('Objective',       obj_vars, 'objectives'),
        ('Design variable', dv_vars,  'design_vars'),
        ('Constraint',      con_vars, 'constraints'),
    ]:
        for var_name in var_list:
            pretty, unit = _get_display(var_name)

            # Peek at the final-iteration shape from the first run that has it
            sample_data = None
            for lab in run_labels:
                d = all_histories[lab][cat_key].get(var_name)
                if d is not None:
                    sample_data = d
                    break

            if sample_data is None:
                continue

            n_components = sample_data.shape[1] if sample_data.ndim == 2 else 1
            is_vector = (n_components > 1)

            if is_vector and vector_mode == 'components':
                # One row per component
                for comp in range(n_components):
                    row_values = {}
                    for lab in run_labels:
                        data = all_histories[lab][cat_key].get(var_name)
                        if data is not None:
                            final_val = data[-1]  # last iteration
                            row_values[lab] = final_val[comp]
                        else:
                            row_values[lab] = None
                    rows.append({
                        'Category':      cat_label,
                        'Variable':      f"{pretty} [{comp}]",
                        'OpenMDAO path': f"{var_name}[{comp}]",
                        'Unit':          unit,
                        **row_values,
                    })
            else:
                # Scalar, or vector reduced to a single value
                row_values = {}
                for lab in run_labels:
                    data = all_histories[lab][cat_key].get(var_name)
                    if data is not None:
                        final_val = data[-1]
                        if is_vector:
                            row_values[lab] = _reduce(final_val, vector_mode)
                        else:
                            row_values[lab] = float(final_val.flat[0])
                    else:
                        row_values[lab] = None

                suffix = f" ({vector_mode})" if is_vector else ""
                rows.append({
                    'Category':      cat_label,
                    'Variable':      f"{pretty}{suffix}",
                    'OpenMDAO path': var_name,
                    'Unit':          unit,
                    **row_values,
                })

    # ── 4. Build DataFrame and add delta columns ──────────────────────
    col_order = ['Category', 'Variable', 'OpenMDAO path', 'Unit'] + run_labels
    df = pd.DataFrame(rows, columns=col_order)

    # ── 4b. Compute deltas between paired runs ─────────────────────────
    # Each entry: (reference_run, comparison_run)
    # Delta = value(comparison) - value(reference)
    # Delta % = 100 * (value(comparison) - value(reference)) / value(reference)
    delta_pairs = [
        ('A_ptfm', 'A_ptfm_twr'),
        ('B_ptfm', 'B_ptfm_twr'),
        ('C_ptfm', 'C_ptfm_twr'),
    ]

    for ref_run, cmp_run in delta_pairs:
        if ref_run not in df.columns or cmp_run not in df.columns:
            print(f"  WARNING: skipping delta {ref_run} → {cmp_run} "
                  f"(one or both columns missing)")
            continue

        col_abs = f"Δ({cmp_run}−{ref_run})"
        col_pct = f"Δ%({cmp_run}−{ref_run})"

        df[col_abs] = df[cmp_run] - df[ref_run]
        df[col_pct] = df.apply(
            lambda row: (
                100.0 * (row[cmp_run] - row[ref_run]) / row[ref_run]
                if row[ref_run] is not None
                   and row[cmp_run] is not None
                   and row[ref_run] != 0
                else None
            ),
            axis=1,
        )

    df.to_excel(output_file, index=False, sheet_name='Final Iteration')
    print(f"\n✓ Spreadsheet written to: {output_file}")
    print(f"  {len(rows)} rows  ×  {len(run_labels)} runs  "
          f"+ {2 * len([p for p in delta_pairs if p[0] in df.columns and p[1] in df.columns])} delta columns\n")

    # ── 5. Optional formatting with openpyxl ───────────────────────────
    if HAS_OPENPYXL:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                                  fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        cat_fills = {
            'Objective':       PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                                           fill_type="solid"),
            'Design variable': PatternFill(start_color="E2EFDA", end_color="E2EFDA",
                                           fill_type="solid"),
            'Constraint':      PatternFill(start_color="FCE4D6", end_color="FCE4D6",
                                           fill_type="solid"),
        }

        # Identify delta column indices for special formatting
        delta_col_indices = set()
        for col_idx in range(1, ws.max_column + 1):
            hdr = ws.cell(row=1, column=col_idx).value or ""
            if hdr.startswith("Δ"):
                delta_col_indices.add(col_idx)
        delta_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                                 fill_type="solid")
        delta_font = Font(bold=True, size=10)

        # Format header row
        delta_header_fill = PatternFill(start_color="BF8F00", end_color="BF8F00",
                                        fill_type="solid")
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
            hdr = cell.value or ""
            if hdr.startswith("Δ"):
                cell.fill = delta_header_fill
            else:
                cell.fill = header_fill

        # Format data rows
        num_fmt = '0.000000'
        for row_idx in range(2, ws.max_row + 1):
            cat_val = ws.cell(row=row_idx, column=1).value
            row_fill = cat_fills.get(cat_val, PatternFill())
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.fill = row_fill
                # Numeric formatting for run columns (col 5 onward)
                if col_idx >= 5 and isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    # Delta-% columns: show 2 decimal places with % sign
                    if col_idx in delta_col_indices:
                        hdr = ws.cell(row=1, column=col_idx).value or ""
                        if "Δ%" in hdr:
                            cell.number_format = '0.00"%"'
                        else:
                            cell.number_format = num_fmt
                        cell.fill = delta_fill
                        cell.font = delta_font
                    else:
                        cell.number_format = num_fmt

        # Auto-fit column widths (approximate)
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Freeze top row
        ws.freeze_panes = 'A2'

        wb.save(output_file)
        print("  (Formatting applied with openpyxl)\n")

    return df


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  RUN                                                                     ║
# ╚═══════════════════════════════════════════════════════════════════════════╝

if __name__ == '__main__':
    export_final_iteration(
        run_dict            = runs,
        select_objectives   = select_objectives,
        select_design_vars  = select_design_vars,
        select_constraints  = select_constraints,
        vector_mode         = vector_mode,
        output_file         = output_file,
    )
